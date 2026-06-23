#!/usr/bin/env python3
"""Gera o ficheiro Excel de casos de teste (jornadas de utilizador) do portal dados.gov.pt.

Cobre as funcionalidades em vigor no portal, em Front Office e Back Office,
categorizadas por secção e nível de permissão (ver backend/docs/permissions.md).
"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Níveis de permissão (linguagem simples; níveis superiores herdam os inferiores)
ANON = "Utilizador não autenticado"
USER = "Utilizador registado / autenticado"
EDITOR = "Membro de organização – Editor"
ORGADMIN = "Membro de organização – Admin"
SYSADMIN = "SuperAdmin / sysadmin"

# (Front Office/Back Office, Secção, Permissão, Jornada de utilizador)
JORNADAS = [
    # ===================== FRONT OFFICE =====================
    ("Front Office", "Página inicial", ANON,
     "Aceder à página inicial e ver os destaques (conjuntos de dados, reutilizações, notícias) e os indicadores do portal."),
    ("Front Office", "Pesquisa global", ANON,
     "Usar a barra de pesquisa do cabeçalho, ver as sugestões automáticas e abrir a página de resultados."),
    ("Front Office", "Pesquisa global", ANON,
     "Na página de resultados de pesquisa, refinar por tipo de conteúdo e abrir um resultado."),

    ("Front Office", "Conjuntos de dados — Listagem", ANON,
     "Listar conjuntos de dados, aplicar filtros (organização, tema, formato, licença) e ordenar os resultados."),
    ("Front Office", "Conjuntos de dados — Detalhe", ANON,
     "Abrir o detalhe de um conjunto de dados e consultar os metadados e a lista de recursos."),
    ("Front Office", "Conjuntos de dados — Detalhe", ANON,
     "Descarregar um recurso de um conjunto de dados e copiar o URL do recurso."),
    ("Front Office", "Conjuntos de dados — Discussões", ANON,
     "Consultar as discussões associadas a um conjunto de dados."),
    ("Front Office", "Conjuntos de dados — Reutilizações", ANON,
     "Consultar as reutilizações associadas a um conjunto de dados."),
    ("Front Office", "Conjuntos de dados — Favoritos", USER,
     "Adicionar e remover um conjunto de dados dos favoritos (seguir/deixar de seguir)."),
    ("Front Office", "Conjuntos de dados — Discussões", USER,
     "Criar uma nova discussão e comentar numa discussão existente de um conjunto de dados."),

    ("Front Office", "Organizações — Listagem", ANON,
     "Listar e pesquisar organizações."),
    ("Front Office", "Organizações — Detalhe", ANON,
     "Abrir o perfil de uma organização e ver os seus conjuntos de dados e reutilizações."),
    ("Front Office", "Organizações — Detalhe", USER,
     "Seguir uma organização a partir do seu perfil."),

    ("Front Office", "Reutilizações — Listagem", ANON,
     "Listar reutilizações e filtrar por tipo e tema."),
    ("Front Office", "Reutilizações — Detalhe", ANON,
     "Abrir o detalhe de uma reutilização e ver os conjuntos de dados utilizados."),
    ("Front Office", "Reutilizações — Detalhe", USER,
     "Seguir uma reutilização e participar na sua discussão."),

    ("Front Office", "Dataservices (APIs)", ANON,
     "Listar dataservices e abrir o detalhe / pré-visualização de um serviço de dados."),
    ("Front Office", "Temas", ANON,
     "Explorar a página de temas e navegar para os conteúdos relacionados com um tema."),
    ("Front Office", "Notícias / Artigos", ANON,
     "Listar notícias e abrir o detalhe de uma notícia."),
    ("Front Office", "Data Stories", ANON,
     "Explorar as data stories e abrir uma história de dados."),
    ("Front Office", "Aprender / Mini-cursos", ANON,
     "Aceder à área de aprendizagem, abrir um mini-curso e navegar pelos seus passos."),
    ("Front Office", "Páginas informativas", ANON,
     "Consultar FAQs, «Sobre dados abertos», Roteiro e Apoio/Contactos."),
    ("Front Office", "Documentação da API", ANON,
     "Consultar a documentação e o tutorial da API."),
    ("Front Office", "Perfil público de utilizador", ANON,
     "Consultar o perfil público de um utilizador e os seus contributos."),

    ("Front Office", "Autenticação — Registo", ANON,
     "Registar uma nova conta e validar o acesso após registo."),
    ("Front Office", "Autenticação — Login", ANON,
     "Autenticar-se com credenciais válidas e ver mensagem de erro com credenciais inválidas."),
    ("Front Office", "Autenticação — Palavra-passe", ANON,
     "Recuperar e redefinir a palavra-passe esquecida através do email."),
    ("Front Office", "Autenticação — Migração de conta", ANON,
     "Migrar uma conta antiga através do fluxo «migrar conta»."),
    ("Front Office", "Autenticação — Logout", USER,
     "Terminar a sessão (logout) e confirmar o regresso ao estado não autenticado."),

    # ===================== BACK OFFICE =====================
    ("Back Office", "Conta — Perfil", USER,
     "Aceder ao backoffice e editar os dados do perfil pessoal."),
    ("Back Office", "Conta — Segurança", USER,
     "Alterar a palavra-passe da conta."),
    ("Back Office", "Conta — Segurança", USER,
     "Alterar o endereço de email da conta."),
    ("Back Office", "Conta — Notificações", USER,
     "Consultar as notificações da conta."),
    ("Back Office", "Conta — Estatísticas", USER,
     "Consultar as estatísticas pessoais da conta."),

    ("Back Office", "Os meus conteúdos — Conjuntos de dados", USER,
     "Criar um novo conjunto de dados, incluindo o upload de recursos."),
    ("Back Office", "Os meus conteúdos — Conjuntos de dados", USER,
     "Editar e eliminar um conjunto de dados próprio e gerir os seus recursos."),
    ("Back Office", "Os meus conteúdos — Reutilizações", USER,
     "Criar, editar e eliminar uma reutilização própria."),
    ("Back Office", "Os meus conteúdos — Recursos comunitários", USER,
     "Criar, editar e eliminar um recurso comunitário próprio."),
    ("Back Office", "Os meus conteúdos — Dataservices", USER,
     "Consultar e gerir os dataservices próprios."),
    ("Back Office", "Organizações — Adesão", USER,
     "Pedir adesão a uma organização e acompanhar o estado do pedido."),
    ("Back Office", "Organizações — Criação", USER,
     "Criar uma nova organização."),

    ("Back Office", "Organização — Conjuntos de dados", EDITOR,
     "Listar, criar e editar conjuntos de dados da organização."),
    ("Back Office", "Organização — Reutilizações", EDITOR,
     "Criar e editar reutilizações da organização."),
    ("Back Office", "Organização — Dataservices / Recursos comunitários", EDITOR,
     "Gerir dataservices e recursos comunitários da organização."),
    ("Back Office", "Organização — Discussões", EDITOR,
     "Consultar as discussões dirigidas à organização."),
    ("Back Office", "Organização — Harvesters", EDITOR,
     "Pré-visualizar (preview) os harvesters da organização."),
    ("Back Office", "Organização — Estatísticas", EDITOR,
     "Consultar as estatísticas da organização."),

    ("Back Office", "Organização — Definições", ORGADMIN,
     "Editar as definições/perfil da organização (nome, descrição, logótipo)."),
    ("Back Office", "Organização — Membros", ORGADMIN,
     "Gerir membros (adicionar, remover, alterar papel) e responder a pedidos de adesão."),
    ("Back Office", "Organização — Harvesters", ORGADMIN,
     "Criar, editar, agendar e executar harvesters da organização e consultar os seus jobs."),

    ("Back Office", "Sistema — Conjuntos de dados", SYSADMIN,
     "Gerir todos os conjuntos de dados do portal (incluindo privados)."),
    ("Back Office", "Sistema — Organizações", SYSADMIN,
     "Gerir todas as organizações do portal."),
    ("Back Office", "Sistema — Utilizadores", SYSADMIN,
     "Gerir todos os utilizadores do portal."),
    ("Back Office", "Sistema — Reutilizações / Dataservices", SYSADMIN,
     "Gerir todas as reutilizações e dataservices do portal."),
    ("Back Office", "Sistema — Harvesters", SYSADMIN,
     "Gerir todos os harvesters do portal."),
    ("Back Office", "Sistema — Notícias / Artigos", SYSADMIN,
     "Criar, editar e publicar notícias e artigos (posts)."),
    ("Back Office", "Sistema — Temas", SYSADMIN,
     "Gerir os temas (topics) do portal."),
    ("Back Office", "Sistema — Editorial", SYSADMIN,
     "Gerir o conteúdo editorial e os destaques da página inicial."),
    ("Back Office", "Sistema — Recursos comunitários", SYSADMIN,
     "Gerir todos os recursos comunitários do portal."),
    ("Back Office", "Sistema — Logs", SYSADMIN,
     "Consultar os logs do sistema."),
]

# ---------------------------------------------------------------- estilos
HEADERS = ["ID", "Front Office/Back Office", "Secção", "Permissão",
           "Jornada de utilizador", "Status", "Data", "Jira"]

NAVY = "1F4E78"
FO_FILL = "DDEBF7"   # azul claro
BO_FILL = "FCE4D6"   # laranja claro
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

wb = Workbook()
ws = wb.active
ws.title = "Casos de Teste"

# Cabeçalho
ws.append(HEADERS)
for col, _ in enumerate(HEADERS, start=1):
    c = ws.cell(row=1, column=col)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = CENTER
    c.border = BORDER

# Linhas de dados
for idx, (fobo, seccao, permissao, jornada) in enumerate(JORNADAS, start=1):
    row = [idx, fobo, seccao, permissao, jornada, "", "", ""]
    ws.append(row)
    r = ws.max_row
    fill = PatternFill("solid", fgColor=FO_FILL if fobo == "Front Office" else BO_FILL)
    for col in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=r, column=col)
        cell.border = BORDER
        cell.alignment = CENTER if col in (1, 2, 6, 7, 8) else WRAP_TOP
        if col in (1, 2):
            cell.fill = fill

# Validação de dados na coluna Status (OK / NOK)
dv = DataValidation(type="list", formula1='"OK,NOK"', allow_blank=True)
dv.error = "Selecione OK ou NOK."
dv.prompt = "Selecione OK ou NOK."
ws.add_data_validation(dv)
dv.add(f"F2:F{ws.max_row}")

# Larguras de coluna
widths = {"A": 6, "B": 22, "C": 38, "D": 30, "E": 70, "F": 10, "G": 14, "H": 16}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:H{ws.max_row}"
ws.sheet_view.showGridLines = False

# ---------------------------------------------------------------- Legenda
leg = wb.create_sheet("Legenda")
leg["A1"] = "Legenda — Casos de Teste dados.gov.pt"
leg["A1"].font = Font(bold=True, size=13, color=NAVY)

leg["A3"] = "Coluna"
leg["B3"] = "Significado"
for c in ("A3", "B3"):
    leg[c].font = HEADER_FONT
    leg[c].fill = HEADER_FILL
    leg[c].alignment = CENTER

legenda = [
    ("ID", "Número sequencial do caso de teste."),
    ("Front Office/Back Office", "Front Office = área pública do portal; Back Office = área de gestão/administração."),
    ("Secção", "Área do portal (nível alto) onde a funcionalidade se enquadra."),
    ("Permissão", "Tipo de utilizador necessário para aceder à funcionalidade (ver níveis abaixo)."),
    ("Jornada de utilizador", "Descrição resumida do percurso a verificar."),
    ("Status", "Resultado do teste: OK (funciona) ou NOK (não funciona)."),
    ("Data", "Data em que o teste foi realizado (a preencher)."),
    ("Jira", "Referência do ticket Jira associado, se existir (a preencher)."),
]
r = 4
for k, v in legenda:
    leg.cell(row=r, column=1, value=k).font = Font(bold=True)
    leg.cell(row=r, column=2, value=v)
    leg.cell(row=r, column=2).alignment = WRAP_TOP
    r += 1

r += 1
leg.cell(row=r, column=1, value="Níveis de permissão (os níveis superiores herdam os inferiores)").font = Font(bold=True, color=NAVY)
r += 1
niveis = [
    (ANON, "Qualquer visitante; apenas leitura de conteúdo público."),
    (USER, "Conta criada e com sessão iniciada; gere os seus próprios conteúdos."),
    (EDITOR, "Membro de organização; edita os conteúdos da organização e pré-visualiza harvesters."),
    (ORGADMIN, "Administra a organização: definições, membros e harvesters."),
    (SYSADMIN, "Administrador do portal; acesso total a todas as funcionalidades."),
]
for k, v in niveis:
    leg.cell(row=r, column=1, value=k).font = Font(bold=True)
    leg.cell(row=r, column=2, value=v)
    leg.cell(row=r, column=2).alignment = WRAP_TOP
    r += 1

leg.column_dimensions["A"].width = 34
leg.column_dimensions["B"].width = 80
leg.sheet_view.showGridLines = False

out = "/home/adbrum/workspace/babel/dadosgov/Casos_de_Teste_dados.gov.pt.xlsx"
wb.save(out)
print(f"OK -> {out}")
print(f"Total de jornadas: {len(JORNADAS)}")
fo = sum(1 for j in JORNADAS if j[0] == 'Front Office')
print(f"Front Office: {fo} | Back Office: {len(JORNADAS) - fo}")
