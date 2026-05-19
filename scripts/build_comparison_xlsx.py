"""Generate a comparative Excel sheet between dados.gov.pt and data.gouv.fr.

Output: /home/adbrum/workspace/babel/dadosgov/comparativo_dadosgov_vs_datagouvfr.xlsx
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

# ---------------------------------------------------------------------------
# Permission vocabulary — labels written so a non-technical reader can grasp
# them at a glance. The Legenda sheet explains each in detail.
# ---------------------------------------------------------------------------
ANON = "Qualquer pessoa (sem conta)"
LOGADO = "Utilizador com conta"
ORG = "Membro da Organização (editor/admin)"
SYS = "Administrador do Portal"

YES = "Sim"
NO = "Não"
PARTIAL = "Parcial"

# ---------------------------------------------------------------------------
# (Bloco, Secção, Detalhe, Permissão, dados.gov?, data.gouv.fr?)
#
# Notes for the non-technical reader:
# - "Front Office"  = páginas públicas do portal (o que o cidadão vê).
# - "Back Office"   = áreas de gestão e administração (/admin).
# - Termos técnicos inevitáveis (CKAN, OAuth2, DCAT, INSPIRE…) são mantidos
#   mas acompanhados de uma explicação curta entre parênteses.
# ---------------------------------------------------------------------------
ROWS = [
    # =====================================================================
    # FRONT OFFICE  (páginas públicas)
    # =====================================================================

    # --- Página Inicial -----------------------------------------------------
    ("Front Office", "Página Inicial", "Barra de pesquisa em grande destaque no topo da página inicial", ANON, YES, YES),
    ("Front Office", "Página Inicial", "Painel com números gerais do portal (conjuntos de dados, ficheiros, organizações, reutilizações e utilizadores)", ANON, YES, YES),
    ("Front Office", "Página Inicial", "Bloco com conjuntos de dados em destaque escolhidos pela equipa", ANON, YES, YES),
    ("Front Office", "Página Inicial", "Bloco com reutilizações em destaque escolhidas pela equipa", ANON, YES, YES),
    ("Front Office", "Página Inicial", "Bloco com últimas notícias/publicações do blog", ANON, YES, YES),
    ("Front Office", "Página Inicial", "Logótipos das principais organizações publicadoras e reutilizadoras", ANON, YES, YES),
    ("Front Office", "Página Inicial", "Bloco com portais temáticos federados (vários sub-portais agregados num só)", ANON, NO, YES),
    ("Front Office", "Página Inicial", "Bloco de 'Histórias de Dados' (data stories) — narrativas visuais e interativas curadas", ANON, YES, NO),

    # --- Navegação no site --------------------------------------------------
    ("Front Office", "Navegação no site", "Menu principal com acesso a Conjuntos de Dados, Organizações, Reutilizações, Temas e Blog", ANON, YES, YES),
    ("Front Office", "Navegação no site", "Botão 'Publicar' no topo, com atalhos para criar conjunto de dados, API, reutilização, importador ou organização", LOGADO, YES, YES),
    ("Front Office", "Navegação no site", "Centro de notificações no topo da página (com contador de mensagens por ler)", LOGADO, YES, YES),
    ("Front Office", "Navegação no site", "Trilho de navegação ('está aqui: Página > Secção > ...') em todas as páginas internas", ANON, YES, YES),
    ("Front Office", "Navegação no site", "Rodapé com inscrição na newsletter, ligações às redes sociais e feeds RSS", ANON, YES, YES),
    ("Front Office", "Navegação no site", "Rodapé com ligações para portais temáticos especializados", ANON, NO, YES),

    # --- Pesquisa no portal -------------------------------------------------
    ("Front Office", "Pesquisa no portal", "Pesquisa rápida no topo do site com sugestões automáticas (autocomplete)", ANON, YES, YES),
    ("Front Office", "Pesquisa no portal", "Página única de pesquisa com botão para alternar entre Conjuntos de Dados / APIs / Reutilizações / Organizações", ANON, PARTIAL, YES),
    ("Front Office", "Pesquisa no portal", "Sugestões automáticas enquanto se escreve (autocomplete via /datasets/suggest)", ANON, YES, YES),
    ("Front Office", "Pesquisa no portal", "Motor de pesquisa avançado baseado em Elasticsearch (rápido e tolerante a erros de escrita)", ANON, YES, YES),
    ("Front Office", "Pesquisa no portal", "Painel lateral de filtros avançados, que pode ser recolhido para ganhar espaço", ANON, YES, YES),
    ("Front Office", "Pesquisa no portal", "Páginas com filtros ativos pedem ao Google para não as indexar (boa prática de SEO)", ANON, NO, YES),

    # --- Conjuntos de Dados — listagem e filtros ----------------------------
    ("Front Office", "Conjuntos de Dados", "Lista de conjuntos de dados, com paginação", ANON, YES, YES),
    ("Front Office", "Conjuntos de Dados", "Filtrar por etiqueta (tag) ou categoria", ANON, YES, YES),
    ("Front Office", "Conjuntos de Dados", "Filtrar pela organização que publicou", ANON, YES, YES),
    ("Front Office", "Conjuntos de Dados", "Filtrar pelo formato do ficheiro (CSV, JSON, PDF, etc.)", ANON, YES, YES),
    ("Front Office", "Conjuntos de Dados", "Filtrar pela licença de utilização", ANON, YES, YES),
    ("Front Office", "Conjuntos de Dados", "Filtrar pela frequência de atualização dos dados", ANON, YES, YES),
    ("Front Office", "Conjuntos de Dados", "Filtrar pela área geográfica abrangida (país, distrito, freguesia…)", ANON, PARTIAL, YES),
    ("Front Office", "Conjuntos de Dados", "Filtrar pelo nível de detalhe geográfico (nacional, regional, local…)", ANON, PARTIAL, YES),
    ("Front Office", "Conjuntos de Dados", "Filtrar pelo esquema/estrutura dos dados (TableSchema para ficheiros tabulares)", ANON, NO, YES),
    ("Front Office", "Conjuntos de Dados", "Filtrar pelo tipo de entidade produtora (pública/privada/associativa…)", ANON, NO, YES),
    ("Front Office", "Conjuntos de Dados", "Filtrar por distinções (badges) da organização publicadora", ANON, NO, YES),
    ("Front Office", "Conjuntos de Dados", "Filtrar pelo intervalo de datas da última atualização", ANON, NO, YES),
    ("Front Office", "Conjuntos de Dados", "Ordenar resultados (mais recentes, mais relevantes, mais seguidos, etc.)", ANON, YES, YES),

    # --- Conjuntos de Dados — página de detalhe -----------------------------
    ("Front Office", "Conjuntos de Dados", "Página de detalhe com descrição em texto formatado (Markdown)", ANON, YES, YES),
    ("Front Office", "Conjuntos de Dados", "Barra/indicador de qualidade do conjunto de dados", ANON, YES, YES),
    ("Front Office", "Conjuntos de Dados", "Distintivos (badges) do conjunto de dados — ex.: HVD (dados de alto valor), INSPIRE (geográficos), SPD (dados de referência), Spam, etc.", ANON, YES, YES),
    ("Front Office", "Conjuntos de Dados", "Separadores na página: Ficheiros / Informações / Reutilizações / Discussões", ANON, YES, YES),
    ("Front Office", "Conjuntos de Dados", "Lista de ficheiros agrupados por tipo (principais, documentação, etc.)", ANON, YES, YES),
    ("Front Office", "Conjuntos de Dados", "Descarregar diretamente os ficheiros associados", ANON, YES, YES),
    ("Front Office", "Conjuntos de Dados", "Pré-visualização de ficheiros CSV/tabela diretamente no browser", ANON, PARTIAL, YES),
    ("Front Office", "Conjuntos de Dados", "Pré-visualização de ficheiros JSON / XML no browser", ANON, NO, YES),
    ("Front Office", "Conjuntos de Dados", "Pré-visualização de ficheiros PDF no browser", ANON, NO, YES),
    ("Front Office", "Conjuntos de Dados", "Pré-visualização de mapas (GeoJSON/PMTiles) com Leaflet (mapas interativos)", ANON, PARTIAL, YES),
    ("Front Office", "Conjuntos de Dados", "Visualizador avançado Datafair (filtros, métricas e mapas enriquecidos)", ANON, NO, YES),
    ("Front Office", "Conjuntos de Dados", "Pré-visualização de tabelas Grist (folhas de cálculo colaborativas, do numerique.gouv.fr)", ANON, NO, YES),
    ("Front Office", "Conjuntos de Dados", "Distintivo de esquema + ficha técnica da estrutura dos dados (schema.data.gouv.fr)", ANON, NO, YES),
    ("Front Office", "Conjuntos de Dados", "Caixa com estatísticas de visualizações e descarregamentos (StatBox)", ANON, YES, YES),
    ("Front Office", "Conjuntos de Dados", "Sugestões de conjuntos de dados e reutilizações relacionados", ANON, NO, YES),
    ("Front Office", "Conjuntos de Dados", "Código para incorporar o conjunto noutro site (snippet copiável, oEmbed)", ANON, PARTIAL, YES),
    ("Front Office", "Conjuntos de Dados", "Partilha social — copiar ligação e partilhar em redes sociais", ANON, YES, YES),
    ("Front Office", "Conjuntos de Dados", "Botão 'Seguir' para receber alertas sobre o conjunto de dados", LOGADO, YES, YES),
    ("Front Office", "Conjuntos de Dados", "Botão 'Denunciar' (spam ou conteúdo impróprio)", LOGADO, NO, YES),
    ("Front Office", "Conjuntos de Dados", "Exportar metadados em formatos abertos RDF/DCAT (XML, Turtle, JSON-LD, N3)", ANON, YES, YES),
    ("Front Office", "Conjuntos de Dados", "Feed Atom/RSS com os conjuntos de dados mais recentes (/datasets/recent.atom)", ANON, YES, YES),
    ("Front Office", "Conjuntos de Dados", "Dados estruturados JSON-LD (Schema.org) — ajudam o Google a perceber o conteúdo", ANON, PARTIAL, YES),
    ("Front Office", "Conjuntos de Dados", "Imagem de pré-visualização (OG-image) gerada automaticamente para partilha em redes sociais", ANON, NO, YES),

    # --- Recursos da Comunidade --------------------------------------------
    ("Front Office", "Conjuntos de Dados", "Separador 'Recursos da Comunidade' — ficheiros enviados por outros utilizadores sobre o mesmo conjunto", ANON, YES, YES),
    ("Front Office", "Conjuntos de Dados", "Submeter um ficheiro próprio como 'recurso da comunidade' num conjunto de dados existente", LOGADO, YES, YES),

    # --- Reutilizações -----------------------------------------------------
    ("Front Office", "Reutilizações (casos de uso)", "Lista de reutilizações (aplicações/estudos/visualizações que usam os dados), com paginação", ANON, YES, YES),
    ("Front Office", "Reutilizações (casos de uso)", "Filtros por tipo, organização, distintivos (badges) e etiqueta (tag)", ANON, YES, YES),
    ("Front Office", "Reutilizações (casos de uso)", "Página de detalhe da reutilização (imagem, descrição, conjuntos de dados e APIs usados)", ANON, YES, YES),
    ("Front Office", "Reutilizações (casos de uso)", "Submeter uma nova reutilização — assistente passo-a-passo (3 etapas)", LOGADO, YES, YES),
    ("Front Office", "Reutilizações (casos de uso)", "Botão 'Seguir' a reutilização para receber alertas", LOGADO, YES, YES),
    ("Front Office", "Reutilizações (casos de uso)", "Feed Atom com as reutilizações mais recentes", ANON, YES, YES),
    ("Front Office", "Reutilizações (casos de uso)", "Discussões e comentários na página da reutilização", ANON, YES, YES),

    # --- Organizações ------------------------------------------------------
    ("Front Office", "Organizações Publicadoras", "Lista de organizações com pesquisa", ANON, YES, YES),
    ("Front Office", "Organizações Publicadoras", "Página da organização com cabeçalho (logótipo, descrição, distintivos)", ANON, YES, YES),
    ("Front Office", "Organizações Publicadoras", "Separadores: Conjuntos de Dados / APIs / Reutilizações / Informações", ANON, YES, YES),
    ("Front Office", "Organizações Publicadoras", "Sistema de distintivos (badges) da organização — ex.: pública, certificada, etc.", ANON, YES, YES),
    ("Front Office", "Organizações Publicadoras", "Selo 'Service-Public' — entidade pública verificada pelo Estado", ANON, NO, YES),
    ("Front Office", "Organizações Publicadoras", "Pesquisa dentro do conteúdo da organização", ANON, YES, YES),
    ("Front Office", "Organizações Publicadoras", "Botão 'Seguir' a organização", LOGADO, YES, YES),
    ("Front Office", "Organizações Publicadoras", "Pedir para se juntar à organização (pedido de adesão)", LOGADO, YES, YES),
    ("Front Office", "Organizações Publicadoras", "Exportar para CSV a lista de conjuntos de dados / reutilizações / discussões da organização", ANON, YES, YES),
    ("Front Office", "Organizações Publicadoras", "Catálogo aberto da organização em RDF/DCAT (padrão europeu de metadados)", ANON, YES, YES),

    # --- APIs / Serviços de Dados ------------------------------------------
    ("Front Office", "Serviços de Dados (APIs)", "Catálogo público de APIs disponíveis", ANON, NO, YES),
    ("Front Office", "Serviços de Dados (APIs)", "Página de cada API com descrição e conjuntos de dados associados", ANON, NO, YES),
    ("Front Office", "Serviços de Dados (APIs)", "Documentação interativa OpenAPI integrada (Swagger UI)", ANON, NO, YES),
    ("Front Office", "Serviços de Dados (APIs)", "Indicação do tipo de acesso (aberto / restrito) e quem pode usar a API", ANON, NO, YES),
    ("Front Office", "Serviços de Dados (APIs)", "Discussões e comentários na página da API", ANON, NO, YES),
    ("Front Office", "Serviços de Dados (APIs)", "Submeter uma nova API por parte do utilizador", LOGADO, NO, YES),

    # --- Temas / Categorias -------------------------------------------------
    ("Front Office", "Temas / Categorias", "Página com a lista de temas e categorias", ANON, YES, YES),
    ("Front Office", "Temas / Categorias", "Página de cada tema com conjuntos de dados e reutilizações selecionados", ANON, YES, YES),
    ("Front Office", "Temas / Categorias", "Página temática sazonal dedicada (ex.: 'Eleições', 'Verão de incêndios')", ANON, NO, YES),

    # --- Blog / Notícias ----------------------------------------------------
    ("Front Office", "Blog / Notícias", "Lista de publicações com paginação", ANON, YES, YES),
    ("Front Office", "Blog / Notícias", "Página de detalhe da publicação (texto Markdown + blocos editoriais)", ANON, YES, YES),
    ("Front Office", "Blog / Notícias", "Publicações em modo 'página editorial completa' (layout livre por blocos)", ANON, NO, YES),

    # --- Discussões e Comentários ------------------------------------------
    ("Front Office", "Discussões e Comentários", "Tópicos de discussão em conjuntos de dados, reutilizações e APIs", ANON, YES, YES),
    ("Front Office", "Discussões e Comentários", "Submeter um comentário ou abrir uma discussão", LOGADO, YES, YES),
    ("Front Office", "Discussões e Comentários", "Editar ou apagar comentários da própria autoria", LOGADO, YES, YES),
    ("Front Office", "Discussões e Comentários", "Denunciar spam num comentário ou numa discussão", LOGADO, NO, YES),

    # --- Início de Sessão / Registo ----------------------------------------
    ("Front Office", "Início de Sessão / Registo", "Entrar com email e palavra-passe", ANON, YES, YES),
    ("Front Office", "Início de Sessão / Registo", "Criar nova conta (registo)", ANON, YES, YES),
    ("Front Office", "Início de Sessão / Registo", "Recuperar palavra-passe esquecida (por email)", ANON, YES, YES),
    ("Front Office", "Início de Sessão / Registo", "Confirmação do endereço de email no registo", ANON, YES, YES),
    ("Front Office", "Início de Sessão / Registo", "Entrar com credenciais do Estado (Autenticação.gov em Portugal / ProConnect em França)", ANON, YES, YES),
    ("Front Office", "Início de Sessão / Registo", "Entrar com sistemas externos via SAML/OAuth2 (login federado)", ANON, YES, YES),
    ("Front Office", "Início de Sessão / Registo", "Autenticação em 2 passos (2FA) com aplicação geradora de códigos (TOTP) e configuração por QR code", LOGADO, NO, YES),
    ("Front Office", "Início de Sessão / Registo", "Página de consentimento OAuth2 — autorizar uma aplicação externa a aceder à conta", LOGADO, YES, YES),
    ("Front Office", "Início de Sessão / Registo", "CAPTCHA estatal francês (Captchetat) para evitar registos automáticos", ANON, NO, YES),
    ("Front Office", "Início de Sessão / Registo", "Migração de contas antigas do portal anterior", ANON, YES, NO),

    # --- Perfil Público -----------------------------------------------------
    ("Front Office", "Perfil Público", "Página pública do utilizador com métricas e conteúdos publicados", ANON, YES, YES),
    ("Front Office", "Perfil Público", "Seguir um utilizador para receber alertas sobre a sua atividade", LOGADO, YES, YES),

    # --- Páginas Institucionais --------------------------------------------
    ("Front Office", "Páginas Institucionais", "Página 'Sobre' / 'Quem somos'", ANON, YES, YES),
    ("Front Office", "Páginas Institucionais", "FAQ — Perguntas Frequentes / Documentação de ajuda", ANON, YES, YES),
    ("Front Office", "Páginas Institucionais", "Página de boas-vindas para quem publica dados (produtores)", ANON, YES, YES),
    ("Front Office", "Páginas Institucionais", "Página de boas-vindas para quem reutiliza dados", ANON, NO, YES),
    ("Front Office", "Páginas Institucionais", "Termos e condições de utilização", ANON, YES, YES),
    ("Front Office", "Páginas Institucionais", "Política de privacidade e proteção de dados (RGPD)", ANON, YES, YES),
    ("Front Office", "Páginas Institucionais", "Declaração de acessibilidade (cumprimento RGAA/WCAG)", ANON, YES, YES),
    ("Front Office", "Páginas Institucionais", "Formulário de contacto", ANON, YES, YES),
    ("Front Office", "Páginas Institucionais", "Painel público de estatísticas com gráficos dos últimos 12 meses", ANON, NO, YES),
    ("Front Office", "Páginas Institucionais", "Catálogo de mini-cursos / área 'Aprender'", ANON, YES, NO),
    ("Front Office", "Páginas Institucionais", "Catálogo de produtos relacionados (ex.: explorador de empresas)", ANON, NO, YES),
    ("Front Office", "Páginas Institucionais", "Página de eventos sazonais / campanhas temporárias", ANON, NO, YES),

    # --- Incorporar conteúdo (Embeds) --------------------------------------
    ("Front Office", "Incorporar conteúdo (Embeds)", "Incorporar um Conjunto de Dados noutro site (embed)", ANON, YES, YES),
    ("Front Office", "Incorporar conteúdo (Embeds)", "Incorporar uma Reutilização noutro site (embed)", ANON, PARTIAL, YES),
    ("Front Office", "Incorporar conteúdo (Embeds)", "Incorporar uma API/Dataservice noutro site (embed)", ANON, NO, YES),
    ("Front Office", "Incorporar conteúdo (Embeds)", "Incorporar uma Organização noutro site (embed)", ANON, NO, YES),
    ("Front Office", "Incorporar conteúdo (Embeds)", "Incorporar uma lista/pesquisa de conjuntos de dados noutro site (embed)", ANON, NO, YES),
    ("Front Office", "Incorporar conteúdo (Embeds)", "Endpoint oEmbed (compatível com WordPress, Medium e outros)", ANON, NO, YES),

    # --- Idiomas, Acessibilidade e SEO -------------------------------------
    ("Front Office", "Idiomas, Acessibilidade e SEO", "Suporte multi-idioma (interface traduzida)", ANON, YES, YES),
    ("Front Office", "Idiomas, Acessibilidade e SEO", "Mapa do site gerado automaticamente para motores de pesquisa (sitemap)", ANON, YES, YES),
    ("Front Office", "Idiomas, Acessibilidade e SEO", "Estatísticas de visitas (Matomo) com opção de desativar por página", ANON, PARTIAL, YES),
    ("Front Office", "Idiomas, Acessibilidade e SEO", "Sistema de design oficial do Estado (DSFR em França / Agora-DS em Portugal)", ANON, YES, YES),

    # --- API Pública --------------------------------------------------------
    ("Front Office", "API Pública (acesso programático)", "Documentação interativa da API (Swagger / OpenAPI)", ANON, YES, YES),
    ("Front Office", "API Pública (acesso programático)", "API REST versão 1 (/api/1/*) — para integradores externos", ANON, YES, YES),
    ("Front Office", "API Pública (acesso programático)", "API REST versão 2 (/api/2/*) — usada para Temas", ANON, YES, YES),
    ("Front Office", "API Pública (acesso programático)", "Catálogo CSV completo do portal disponível para descarregar", ANON, YES, YES),
    ("Front Office", "API Pública (acesso programático)", "Catálogo global do portal em RDF/DCAT-AP (padrão europeu de catálogo aberto de dados)", ANON, YES, YES),
    ("Front Office", "API Pública (acesso programático)", "Endpoint para verificar se um URL de recurso ainda está acessível (check_url)", ANON, YES, YES),

    # =====================================================================
    # BACK OFFICE  (áreas de gestão / administração — /admin)
    # =====================================================================

    # --- Área de Administração ---------------------------------------------
    ("Back Office", "Área de Administração", "Layout do /admin com menu lateral contextual à organização do utilizador", LOGADO, YES, YES),
    ("Back Office", "Área de Administração", "Banner 'modo só leitura' / 'transferência de propriedade pendente'", LOGADO, NO, YES),
    ("Back Office", "Área de Administração", "Proteção de acesso (visitante / autenticado / administrador) ao nível do servidor", LOGADO, YES, YES),

    # --- Minha Conta -------------------------------------------------------
    ("Back Office", "Minha Conta", "Editar perfil próprio (nome, biografia, website)", LOGADO, YES, YES),
    ("Back Office", "Minha Conta", "Carregar ou remover fotografia de perfil (avatar)", LOGADO, YES, YES),
    ("Back Office", "Minha Conta", "Alterar email da conta (com verificação por email)", LOGADO, YES, YES),
    ("Back Office", "Minha Conta", "Alterar palavra-passe", LOGADO, YES, YES),
    ("Back Office", "Minha Conta", "Apagar conta — os dados pessoais ficam anonimizados", LOGADO, YES, YES),
    ("Back Office", "Minha Conta", "Gerir chaves de acesso à API (criar, revogar, ver origem e data de última utilização)", LOGADO, YES, YES),
    ("Back Office", "Minha Conta", "Ativar ou desativar autenticação em 2 passos (2FA com aplicação TOTP)", LOGADO, NO, YES),
    ("Back Office", "Minha Conta", "Listar os meus conjuntos de dados, reutilizações, APIs e recursos da comunidade", LOGADO, YES, YES),
    ("Back Office", "Minha Conta", "Painéis com estatísticas pessoais", LOGADO, YES, YES),
    ("Back Office", "Minha Conta", "Histórico das minhas atividades", LOGADO, YES, YES),
    ("Back Office", "Minha Conta", "Pedidos de transferência de propriedade pendentes de aprovação", LOGADO, YES, YES),
    ("Back Office", "Minha Conta", "Convites recebidos para entrar em organizações (aceitar/recusar)", LOGADO, YES, YES),
    ("Back Office", "Minha Conta", "Lista de utilizadores, organizações, conjuntos e reutilizações que sigo", LOGADO, YES, YES),

    # --- Gestão de Conjuntos de Dados --------------------------------------
    ("Back Office", "Gestão de Conjuntos de Dados", "Lista de conjuntos de dados da organização com filtros de administração", ORG, YES, YES),
    ("Back Office", "Gestão de Conjuntos de Dados", "Criar novo conjunto de dados — assistente passo-a-passo", LOGADO, YES, YES),
    ("Back Office", "Gestão de Conjuntos de Dados", "Assistente de publicação estruturada (com associação a schema.data.gouv.fr)", LOGADO, NO, YES),
    ("Back Office", "Gestão de Conjuntos de Dados", "Editar metadados (título, descrição em Markdown, etiquetas, licença)", ORG, YES, YES),
    ("Back Office", "Gestão de Conjuntos de Dados", "Gerir ficheiros/recursos (adicionar, editar, remover)", ORG, YES, YES),
    ("Back Office", "Gestão de Conjuntos de Dados", "Carregar ficheiros com 'arrastar e largar' (drag-and-drop)", ORG, YES, YES),
    ("Back Office", "Gestão de Conjuntos de Dados", "Substituir um ficheiro existente por uma nova versão", ORG, YES, YES),
    ("Back Office", "Gestão de Conjuntos de Dados", "Editor de tabelas CSV diretamente no navegador (client-side)", ORG, NO, YES),
    ("Back Office", "Gestão de Conjuntos de Dados", "Publicar / despublicar (alternar privado / rascunho / público)", ORG, YES, YES),
    ("Back Office", "Gestão de Conjuntos de Dados", "Arquivar conjunto de dados (deixa de aparecer mas mantém o histórico)", ORG, YES, YES),
    ("Back Office", "Gestão de Conjuntos de Dados", "Eliminar conjunto de dados — eliminação reversível, com possibilidade de restaurar", ORG, YES, YES),
    ("Back Office", "Gestão de Conjuntos de Dados", "Transferir a propriedade para outro utilizador ou organização", ORG, YES, YES),
    ("Back Office", "Gestão de Conjuntos de Dados", "Definir a área geográfica e o período temporal abrangidos", ORG, YES, YES),
    ("Back Office", "Gestão de Conjuntos de Dados", "Associar um esquema de dados (TableSchema) para validar a estrutura dos ficheiros", ORG, NO, YES),
    ("Back Office", "Gestão de Conjuntos de Dados", "Marcar como 'em destaque' (aparece na página inicial)", SYS, YES, YES),
    ("Back Office", "Gestão de Conjuntos de Dados", "Atribuir distintivos (badges) — HVD (alto valor), INSPIRE (geo), SPD (dados de referência), etc.", SYS, YES, YES),
    ("Back Office", "Gestão de Conjuntos de Dados", "Gerar descrição e etiquetas automaticamente com IA (modelo Albert)", ORG, NO, YES),
    ("Back Office", "Gestão de Conjuntos de Dados", "Histórico de atividades do conjunto de dados", ORG, YES, YES),

    # --- Gestão de Organizações --------------------------------------------
    ("Back Office", "Gestão de Organizações", "Criar nova organização — assistente passo-a-passo", LOGADO, YES, YES),
    ("Back Office", "Gestão de Organizações", "Editar perfil (logótipo, nome, descrição, tipo)", ORG, YES, YES),
    ("Back Office", "Gestão de Organizações", "Listar membros da organização e os respetivos papéis", ORG, YES, YES),
    ("Back Office", "Gestão de Organizações", "Convidar um novo membro por email", ORG, YES, YES),
    ("Back Office", "Gestão de Organizações", "Aceitar ou recusar pedidos para entrar na organização", ORG, YES, YES),
    ("Back Office", "Gestão de Organizações", "Alterar o papel de um membro (administrador / editor)", ORG, YES, YES),
    ("Back Office", "Gestão de Organizações", "Remover um membro da organização", ORG, YES, YES),
    ("Back Office", "Gestão de Organizações", "Gerir pontos de contacto da organização", ORG, YES, YES),
    ("Back Office", "Gestão de Organizações", "Painel com estatísticas da organização", ORG, YES, YES),
    ("Back Office", "Gestão de Organizações", "Histórico de atividades da organização", ORG, YES, YES),
    ("Back Office", "Gestão de Organizações", "Listar os importadores automáticos (harvesters) da organização", ORG, YES, YES),
    ("Back Office", "Gestão de Organizações", "Eliminar a organização (eliminação reversível)", ORG, YES, YES),
    ("Back Office", "Gestão de Organizações", "Atribuir distintivos à organização (incluindo Service-Public/Certified)", SYS, YES, YES),

    # --- Gestão de Reutilizações -------------------------------------------
    ("Back Office", "Gestão de Reutilizações", "Lista de reutilizações da organização (vista administrativa)", ORG, YES, YES),
    ("Back Office", "Gestão de Reutilizações", "Editar metadados da reutilização", ORG, YES, YES),
    ("Back Office", "Gestão de Reutilizações", "Associar/desassociar conjuntos de dados e APIs usados", ORG, YES, YES),
    ("Back Office", "Gestão de Reutilizações", "Carregar imagem ilustrativa da reutilização", ORG, YES, YES),
    ("Back Office", "Gestão de Reutilizações", "Acompanhar discussões e atividades da reutilização", ORG, YES, YES),
    ("Back Office", "Gestão de Reutilizações", "Marcar como 'em destaque' (aparece na página inicial)", SYS, YES, YES),
    ("Back Office", "Gestão de Reutilizações", "Atribuir distintivos (badges)", SYS, YES, YES),
    ("Back Office", "Gestão de Reutilizações", "Gerar etiquetas automaticamente com IA (modelo Albert)", ORG, NO, YES),

    # --- Gestão de APIs ----------------------------------------------------
    ("Back Office", "Gestão de APIs", "Criar nova API/Dataservice — assistente passo-a-passo", LOGADO, PARTIAL, YES),
    ("Back Office", "Gestão de APIs", "Editar metadados de uma API", ORG, PARTIAL, YES),
    ("Back Office", "Gestão de APIs", "Associar conjuntos de dados à API", ORG, PARTIAL, YES),
    ("Back Office", "Gestão de APIs", "Configurar tipo de acesso e audiência (aberto / restrito a determinados utilizadores)", ORG, NO, YES),
    ("Back Office", "Gestão de APIs", "Acompanhar discussões e atividades da API", ORG, PARTIAL, YES),

    # --- Gestão de Recursos da Comunidade ----------------------------------
    ("Back Office", "Gestão de Recursos da Comunidade", "Criar um recurso da comunidade", LOGADO, YES, YES),
    ("Back Office", "Gestão de Recursos da Comunidade", "Listar os recursos próprios / da organização", ORG, YES, YES),
    ("Back Office", "Gestão de Recursos da Comunidade", "Eliminar recurso da comunidade", ORG, YES, YES),

    # --- Gestão do Blog ----------------------------------------------------
    ("Back Office", "Gestão do Blog", "Criar/editar publicação em Markdown", SYS, YES, YES),
    ("Back Office", "Gestão do Blog", "Editor por blocos editoriais (cabeçalho hero, listas, acordeão, etc.)", SYS, NO, YES),
    ("Back Office", "Gestão do Blog", "Reordenar blocos editoriais com 'arrastar e largar'", SYS, NO, YES),
    ("Back Office", "Gestão do Blog", "Publicar / despublicar / agendar publicações", SYS, YES, YES),
    ("Back Office", "Gestão do Blog", "Carregar imagem de cabeçalho (hero) da publicação", SYS, YES, YES),

    # --- Gestão de Temas ---------------------------------------------------
    ("Back Office", "Gestão de Temas", "Criar/editar um tema e associar-lhe conjuntos de dados e reutilizações", SYS, PARTIAL, YES),
    ("Back Office", "Gestão de Temas", "Lista global de todos os temas", SYS, PARTIAL, YES),

    # --- Importação automática (Harvesting) --------------------------------
    ("Back Office", "Importação Automática (Harvesting)", "Criar importador automático ('harvester') — assistente com pré-visualização", LOGADO, YES, YES),
    ("Back Office", "Importação Automática (Harvesting)", "Suporte ao protocolo DCAT (catálogo de dados em RDF)", LOGADO, YES, YES),
    ("Back Office", "Importação Automática (Harvesting)", "Suporte a CKAN / CKAN-DCAT (importar de portais CKAN)", LOGADO, YES, YES),
    ("Back Office", "Importação Automática (Harvesting)", "Suporte a OAI-PMH (protocolo de partilha de metadados, comum em bibliotecas/arquivos)", LOGADO, YES, YES),
    ("Back Office", "Importação Automática (Harvesting)", "Suporte a CSW (catálogos geográficos do consórcio OGC)", LOGADO, YES, YES),
    ("Back Office", "Importação Automática (Harvesting)", "Suporte a GeoNetwork (catálogo geográfico) / ODS / RDF", LOGADO, PARTIAL, YES),
    ("Back Office", "Importação Automática (Harvesting)", "Filtros de inclusão/exclusão por etiqueta ou organização (whitelist/blacklist)", LOGADO, YES, YES),
    ("Back Office", "Importação Automática (Harvesting)", "Fluxo de validação pelo Administrador do Portal antes de ativar o importador", SYS, YES, YES),
    ("Back Office", "Importação Automática (Harvesting)", "Agendar execução periódica do importador (a horas/dias fixos)", ORG, YES, YES),
    ("Back Office", "Importação Automática (Harvesting)", "Executar o importador manualmente, sob comando", ORG, YES, YES),
    ("Back Office", "Importação Automática (Harvesting)", "Lista de execuções com estado e erros encontrados", ORG, YES, YES),
    ("Back Office", "Importação Automática (Harvesting)", "Detalhe de cada execução (conjuntos de dados criados/atualizados)", ORG, YES, YES),
    ("Back Office", "Importação Automática (Harvesting)", "Lista global de todos os importadores do portal (apenas Administrador)", SYS, YES, YES),

    # --- Gestão de Utilizadores --------------------------------------------
    ("Back Office", "Gestão de Utilizadores", "Lista global de utilizadores do portal", SYS, YES, YES),
    ("Back Office", "Gestão de Utilizadores", "Editar dados e papéis de qualquer utilizador", SYS, YES, YES),
    ("Back Office", "Gestão de Utilizadores", "Eliminar utilizador (anonimização dos dados pessoais)", SYS, YES, YES),
    ("Back Office", "Gestão de Utilizadores", "Ver conjuntos de dados, reutilizações, APIs e atividades de qualquer utilizador", SYS, YES, YES),
    ("Back Office", "Gestão de Utilizadores", "Atribuir/remover o papel de Administrador do Portal a um utilizador", SYS, YES, YES),

    # --- Moderação ---------------------------------------------------------
    ("Back Office", "Moderação de Conteúdo", "Submeter denúncia (report) sobre conteúdo impróprio", LOGADO, NO, YES),
    ("Back Office", "Moderação de Conteúdo", "Lista de denúncias por tratar", SYS, YES, YES),
    ("Back Office", "Moderação de Conteúdo", "Resolver denúncia (marcar como tratada, apagar o conteúdo)", SYS, YES, YES),
    ("Back Office", "Moderação de Conteúdo", "Página unificada de moderação por tipo de conteúdo denunciado", SYS, NO, YES),
    ("Back Office", "Moderação de Conteúdo", "Marcação de spam (manual e automática)", SYS, YES, YES),
    ("Back Office", "Moderação de Conteúdo", "Moderação das discussões dentro de uma organização", ORG, YES, YES),

    # --- Configuração do Portal (SysAdmin) ---------------------------------
    ("Back Office", "Configuração do Portal", "Editar metadados gerais do portal (/api/1/site/) — nome, descrição, etc.", SYS, YES, YES),
    ("Back Office", "Configuração do Portal", "Editar blocos editoriais das páginas de entrada (conjuntos de dados, reutilizações, …)", SYS, NO, YES),
    ("Back Office", "Configuração do Portal", "Editar páginas dinâmicas tipo CMS (/pages/...)", SYS, PARTIAL, YES),
    ("Back Office", "Configuração do Portal", "Consultar registos do sistema (/api/1/site/logs/)", SYS, YES, YES),
    ("Back Office", "Configuração do Portal", "Listas globais cruzadas (conjuntos de dados, organizações, reutilizações, etc.)", SYS, YES, YES),
    ("Back Office", "Configuração do Portal", "Registos de auditoria (quem fez o quê e quando)", SYS, YES, YES),

    # --- OAuth2 / Chaves de API / Webhooks ---------------------------------
    ("Back Office", "Chaves de API e OAuth2", "Servidor OAuth2 — fluxo 'authorization_code' (login federado para aplicações externas)", SYS, YES, YES),
    ("Back Office", "Chaves de API e OAuth2", "Emissão e revogação de tokens (endpoints /oauth/token e /oauth/revoke)", LOGADO, YES, YES),
    ("Back Office", "Chaves de API e OAuth2", "Gestão das aplicações OAuth2 de terceiros autorizadas", SYS, YES, YES),
    ("Back Office", "Chaves de API e OAuth2", "Autenticação alternativa por cabeçalho HTTP X-API-KEY", LOGADO, YES, YES),
    ("Back Office", "Chaves de API e OAuth2", "Webhooks por organização — notificar sistemas externos quando algo muda", SYS, PARTIAL, PARTIAL),

    # --- Notificações ------------------------------------------------------
    ("Back Office", "Notificações", "Notificação: novo comentário numa discussão que sigo", LOGADO, YES, YES),
    ("Back Office", "Notificações", "Notificação: alguém pediu para entrar na minha organização", LOGADO, YES, YES),
    ("Back Office", "Notificações", "Notificação: o meu pedido de adesão foi aceite ou recusado", LOGADO, YES, YES),
    ("Back Office", "Notificações", "Notificação: pedido de transferência de propriedade", LOGADO, YES, YES),
    ("Back Office", "Notificações", "Notificação: há um importador automático (harvester) à espera de validação", SYS, YES, YES),
    ("Back Office", "Notificações", "Notificação: foi-me atribuído um novo distintivo (badge)", LOGADO, PARTIAL, YES),
    ("Back Office", "Notificações", "Marcar notificação como lida", LOGADO, YES, YES),

    # --- Tarefas Agendadas (Jobs/Scheduler) --------------------------------
    ("Back Office", "Tarefas Agendadas (Jobs)", "Listar tarefas agendadas no servidor (Celery)", SYS, YES, YES),
    ("Back Office", "Tarefas Agendadas (Jobs)", "Criar, atualizar ou eliminar tarefas agendadas", SYS, YES, YES),
    ("Back Office", "Tarefas Agendadas (Jobs)", "Ver o resultado/estado de uma tarefa específica", SYS, YES, YES),
    ("Back Office", "Tarefas Agendadas (Jobs)", "Listar todas as tarefas que podem ser agendadas", SYS, YES, YES),

    # --- Dados Geográficos -------------------------------------------------
    ("Back Office", "Dados Geográficos", "Sugestões automáticas (autocomplete) de zonas administrativas — concelhos, distritos…", ANON, YES, YES),
    ("Back Office", "Dados Geográficos", "Listar níveis administrativos e a sua granularidade (nacional → freguesia)", ANON, YES, YES),
    ("Back Office", "Dados Geográficos", "Listar conjuntos de dados associados a uma determinada zona", ANON, YES, YES),
]


# ---------------------------------------------------------------------------
# Build the workbook
# ---------------------------------------------------------------------------
wb = Workbook()
ws = wb.active
ws.title = "Comparativo"

HEADERS = [
    "ID",
    "Front Office/Back Office",
    "Secção",
    "Permissão",
    "Detalhe",
    "Existe no dados.gov.pt?",
    "Existe no data.gouv.fr?",
]

# Styles
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", fgColor="1F4E78")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

front_fill = PatternFill("solid", fgColor="E7F1FA")  # light blue
back_fill = PatternFill("solid", fgColor="FFF2CC")   # light yellow

thin = Side(border_style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

wrap_align = Alignment(vertical="center", wrap_text=True)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Header row
ws.append(HEADERS)
for col_idx, _ in enumerate(HEADERS, start=1):
    cell = ws.cell(row=1, column=col_idx)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = border

# Data rows
# Sort: Front Office block first, then Back Office. Within blocks, keep insertion order.
ordered = [r for r in ROWS if r[0] == "Front Office"] + [r for r in ROWS if r[0] == "Back Office"]

for idx, (bloco, seccao, detalhe, perm, dgov, dgouv) in enumerate(ordered, start=1):
    row = [idx, bloco, seccao, perm, detalhe, dgov, dgouv]
    ws.append(row)
    row_idx = ws.max_row
    fill = front_fill if bloco == "Front Office" else back_fill
    for col_idx in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.fill = fill
        cell.border = border
        cell.alignment = center_align if col_idx in (1, 2, 4, 6, 7) else wrap_align

# Column widths — Permissão widened to fit the longer, friendlier labels.
widths = {"A": 6, "B": 16, "C": 30, "D": 34, "E": 80, "F": 18, "G": 18}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

# Freeze header
ws.freeze_panes = "A2"

# AutoFilter
ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{ws.max_row}"

# Conditional formatting for Sim / Não / Parcial in columns F and G
green_fill = PatternFill("solid", fgColor="C6EFCE")
green_font = Font(color="006100", bold=True)
red_fill = PatternFill("solid", fgColor="FFC7CE")
red_font = Font(color="9C0006", bold=True)
orange_fill = PatternFill("solid", fgColor="FFEB9C")
orange_font = Font(color="9C5700", bold=True)

for col_letter in ("F", "G"):
    rng = f"{col_letter}2:{col_letter}{ws.max_row}"
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Sim"'], fill=green_fill, font=green_font))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Não"'], fill=red_fill, font=red_font))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Parcial"'], fill=orange_fill, font=orange_font))

# ---------------------------------------------------------------------------
# Legend sheet
# ---------------------------------------------------------------------------
ws2 = wb.create_sheet("Legenda")
ws2.append(["Coluna", "Valores possíveis", "Significado"])
for c in range(1, 4):
    cell = ws2.cell(row=1, column=c)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = border

legend_rows = [
    ("Front Office / Back Office", "Front Office", "Páginas públicas do portal, visíveis a qualquer visitante (ex.: lista de conjuntos de dados, página de uma reutilização, etc.)."),
    ("Front Office / Back Office", "Back Office", "Áreas privadas de gestão e administração, acessíveis através de /admin após início de sessão."),
    ("Permissão", "Qualquer pessoa (sem conta)", "Não é necessário iniciar sessão — qualquer visitante consegue aceder."),
    ("Permissão", "Utilizador com conta", "É necessário estar autenticado, mas basta ter uma conta normal no portal."),
    ("Permissão", "Membro da Organização (editor/admin)", "Só é acessível a quem pertence à organização responsável pelo conteúdo, com papel de editor ou de administrador."),
    ("Permissão", "Administrador do Portal", "Reservado a administradores globais do portal (papel 'admin' do udata)."),
    ("Existe no dados.gov.pt? / data.gouv.fr?", "Sim", "Funcionalidade implementada e disponível."),
    ("Existe no dados.gov.pt? / data.gouv.fr?", "Não", "Funcionalidade ausente — ainda não implementada."),
    ("Existe no dados.gov.pt? / data.gouv.fr?", "Parcial", "Existe no motor (udata) mas a interface ainda não a expõe, ou existe uma forma equivalente mas incompleta face ao outro portal."),
]
for r in legend_rows:
    ws2.append(r)
    row_idx = ws2.max_row
    for c in range(1, 4):
        cell = ws2.cell(row=row_idx, column=c)
        cell.alignment = wrap_align
        cell.border = border

ws2.column_dimensions["A"].width = 32
ws2.column_dimensions["B"].width = 36
ws2.column_dimensions["C"].width = 95
ws2.freeze_panes = "A2"

# ---------------------------------------------------------------------------
# Summary sheet
# ---------------------------------------------------------------------------
ws3 = wb.create_sheet("Resumo")
ws3.append(["Métrica", "Valor"])
for c in range(1, 3):
    cell = ws3.cell(row=1, column=c)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = border

total = len(ordered)
front_n = sum(1 for r in ordered if r[0] == "Front Office")
back_n = total - front_n
dgov_yes = sum(1 for r in ordered if r[4] == YES)
dgov_partial = sum(1 for r in ordered if r[4] == PARTIAL)
dgov_no = sum(1 for r in ordered if r[4] == NO)
dgouv_yes = sum(1 for r in ordered if r[5] == YES)
dgouv_partial = sum(1 for r in ordered if r[5] == PARTIAL)
dgouv_no = sum(1 for r in ordered if r[5] == NO)
common = sum(1 for r in ordered if r[4] == YES and r[5] == YES)
only_dgov = sum(1 for r in ordered if r[4] == YES and r[5] == NO)
only_dgouv = sum(1 for r in ordered if r[4] == NO and r[5] == YES)

summary_rows = [
    ("Total de funcionalidades", total),
    ("Front Office", front_n),
    ("Back Office", back_n),
    ("Comuns (Sim em ambos)", common),
    ("Apenas no dados.gov.pt", only_dgov),
    ("Apenas no data.gouv.fr", only_dgouv),
    ("dados.gov.pt - Sim", dgov_yes),
    ("dados.gov.pt - Parcial", dgov_partial),
    ("dados.gov.pt - Não", dgov_no),
    ("data.gouv.fr - Sim", dgouv_yes),
    ("data.gouv.fr - Parcial", dgouv_partial),
    ("data.gouv.fr - Não", dgouv_no),
]
for r in summary_rows:
    ws3.append(r)
    row_idx = ws3.max_row
    for c in range(1, 3):
        cell = ws3.cell(row=row_idx, column=c)
        cell.border = border
        cell.alignment = wrap_align

ws3.column_dimensions["A"].width = 35
ws3.column_dimensions["B"].width = 14
ws3.freeze_panes = "A2"

# ---------------------------------------------------------------------------
output_path = "/home/adbrum/workspace/babel/dadosgov/comparativo_dadosgov_vs_datagouvfr.xlsx"
wb.save(output_path)
print(f"OK -> {output_path}")
print(f"Total rows: {total} (Front: {front_n} / Back: {back_n})")
print(f"Comuns: {common} | Só dados.gov.pt: {only_dgov} | Só data.gouv.fr: {only_dgouv}")
